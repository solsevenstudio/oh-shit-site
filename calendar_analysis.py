import pandas as pd
from ics import Calendar
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference

# === CONFIG ===
ics_file = "Calendar.ics"  # Ensure this file exists in the same folder
excel_file = "calendar_analysis.xlsx"

# === STEP 1: Parse ICS ===
with open(ics_file, 'r', encoding='utf-8') as f:
    calendar = Calendar(f.read())

# Extract events
records = []
for event in calendar.events:
    start = event.begin.datetime
    end = event.end.datetime
    duration_hours = (end - start).total_seconds() / 3600
    records.append({
        "Title": event.name,
        "Start": start,
        "End": end,
        "Duration (hrs)": duration_hours
    })

# Convert to DataFrame
df = pd.DataFrame(records)
df["Start"] = pd.to_datetime(df["Start"], utc=True)
df["End"] = pd.to_datetime(df["End"], utc=True)

# Categorize events by keywords
keywords = {
    "Meeting": ["meeting", "sync", "review"],
    "Call": ["call", "phone", "zoom"],
    "Workshop": ["workshop", "training"],
    "Other": []
}

def categorize(title):
    title_lower = title.lower() if title else ""
    for category, words in keywords.items():
        if any(word in title_lower for word in words):
            return category
    return "Other"

df["Category"] = df["Title"].apply(categorize)
df["Month"] = df["Start"].dt.to_period("M")
df["Week"] = df["Start"].dt.to_period("W")

# Summaries
summary_category = df.groupby("Category")["Duration (hrs)"].sum().sort_values(ascending=False)
summary_month = df.groupby("Month")["Duration (hrs)"].sum()
summary_week = df.groupby("Week")["Duration (hrs)"].sum().sort_values(ascending=False)

# === STEP 2: Create Excel Workbook ===
wb = Workbook()

# Sheet 1: All Events
ws_all = wb.active
ws_all.title = "All Events"
for r in dataframe_to_rows(df, index=False, header=True):
    ws_all.append(r)

# Sheet 2: By Category
ws_cat = wb.create_sheet("By Category")
ws_cat.append(["Category", "Total Hours"])
for cat, hours in summary_category.items():
    ws_cat.append([cat, hours])

# Sheet 3: By Month
ws_month = wb.create_sheet("By Month")
ws_month.append(["Month", "Total Hours"])
for month, hours in summary_month.items():
    ws_month.append([str(month), hours])

# Sheet 4: By Week
ws_week = wb.create_sheet("By Week")
ws_week.append(["Week", "Total Hours"])
for week, hours in summary_week.items():
    ws_week.append([str(week), hours])

# Highlight top 5 busiest weeks
for row in range(2, 7):
    ws_week[f"A{row}"].font = Font(bold=True)
    ws_week[f"B{row}"].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# === STEP 3: Add Charts ===
# Pie chart for Category
pie = PieChart()
pie.title = "Time by Category"
data = Reference(ws_cat, min_col=2, min_row=1, max_row=len(summary_category)+1)
labels = Reference(ws_cat, min_col=1, min_row=2, max_row=len(summary_category)+1)
pie.add_data(data, titles_from_data=True)
pie.set_categories(labels)
ws_cat.add_chart(pie, "D2")

# Bar chart for Month
bar = BarChart()
bar.title = "Hours per Month"
data = Reference(ws_month, min_col=2, min_row=1, max_row=len(summary_month)+1)
labels = Reference(ws_month, min_col=1, min_row=2, max_row=len(summary_month)+1)
bar.add_data(data, titles_from_data=True)
bar.set_categories(labels)
ws_month.add_chart(bar, "D2")

# Save workbook
wb.save(excel_file)
print(f"✅ Excel workbook '{excel_file}' created successfully with charts and highlights!")